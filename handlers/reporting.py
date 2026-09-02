"""
Отчётность — команда /report (send_report).

Исторический R считается только по доказанному риску конкретной сделки из
durable-записи её входа. Текущий глобальный риск знаменателем не является:
изменение /risk обязано влиять на новые сделки, а не переписывать статистику уже
закрытых. Сделка без доказанного риска получает UNKNOWN, а не выдуманный R.

Связать closed-PnL строку с её входом разрешено ровно двумя точными путями:

1. прямой — ``(symbol, orderId)`` строки совпадает с ``(symbol, order_id)``
   сохранённого входа. Так закрывается случай, когда позицию закрыл сам ордер;
2. связь выхода — ``(symbol, orderId)`` строки совпадает с
   ``(symbol, exit_order_id)`` durable-события ``EXIT_ORDER_BOUND``, записанного
   ДО закрытия, пока защитный ордер ещё был виден в открытых ордерах биржи. Так
   закрывается обычный случай Bybit V5, где позицию закрыл дочерний SL/TP.

Второй путь существует потому, что после исполнения защитного ордера биржа
связи с входом не отдаёт вовсе: у SL/TP-детей ``orderLinkId`` и
``parentOrderLinkId`` пустые (доказано production-диагностикой). Поэтому
post-close реконструкция родства невозможна в принципе, и знаменатель обязан
быть сохранён заранее.

Корреляция по символу, времени, цене, объёму, стороне, источнику или близости
записей запрещена: она способна приписать сделке чужой риск. Отсутствие обоих
доказательств — это UNKNOWN, а не выдуманный R.

Страницы закрытых сделок читаются по официальному контракту Bybit V5: токен
продолжения приходит в ``result["nextPageCursor"]`` и уходит следующим запросом
параметром ``cursor``. ``result["cursor"]`` токеном продолжения этого эндпоинта
не является. Неполная выборка страниц агрегатом отчёта стать не может: занижённые
PnL, R, winrate и число сделок выглядят как правда, поэтому любая аномалия
пагинации — ошибка отчёта, а не «данные закончились».
"""

import io
import logging
import math
import asyncio
from datetime import datetime, timedelta, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from telegram import Update
from telegram.ext import ContextTypes

from core.config import ALLOWED_ID
from core.trading_core import session
from core.database import get_source_at_time
from core.journal import (
    UNKNOWN,
    get_entry_risk_evidence,
    get_exit_order_risk_evidence,
    normalize_symbol,
)
from handlers.orders import bybit_call
from handlers.ui import (
    format_action,
    format_bybit_error_detail,
    format_error_message,
    format_header,
    format_value_block,
    h,
)

# Максимально допустимый диапазон одного запроса к Bybit (< 7 суток)
_CHUNK_MS = 7 * 24 * 60 * 60 * 1000 - 1

# Размер страницы Get Closed PnL. Официальный диапазон этого эндпоинта — 1..100;
# лимиты других эндпоинтов (например Get Order History) сюда не переносятся.
_PAGE_LIMIT = 100

# Предел страниц пагинации на один чанк — общий safety bound отчёта: без него
# некорректный cursor Bybit крутил бы цикл бесконечно.
_MAX_PAGES = 50


class _BybitReportError(Exception):
    """Ошибка API Bybit при сборе отчёта."""


def _validate_resp(resp, current_start: int, current_end: int) -> list:
    """
    Проверяет ответ одной страницы get_closed_pnl.

    Возвращает список сделок только при полностью доказанном успехе: ``resp`` —
    dict, ``retCode`` — точно ``int`` со значением ``0``, ``result`` — dict, а
    ``result["list"]`` — list. ``True``/``False`` кодом ответа не являются:
    ``False == 0`` в Python, и без проверки типа ошибочный ответ прошёл бы как
    успешный.

    При любом отклонении поднимает _BybitReportError с деталями (включая
    временное окно чанка) для пользователя и лога: недостоверная страница не
    имеет права попасть в authoritative-агрегат отчёта.
    """
    chunk_info = f"[{current_start}–{current_end}]"
    if not isinstance(resp, dict):
        raise _BybitReportError(
            f"{chunk_info} retCode=—, retMsg=неожиданный тип ответа: {type(resp).__name__}"
        )
    if "retCode" not in resp:
        raise _BybitReportError(
            f"{chunk_info} нет ключа retCode: пустой/невалидный ответ от Bybit; "
            "возможна скрытая ошибка внутри bybit_call"
        )
    ret_code = resp["retCode"]
    if type(ret_code) is not int:
        raise _BybitReportError(
            f"{chunk_info} retCode={ret_code!r}, "
            f"retMsg=недоказанный тип retCode: {type(ret_code).__name__}"
        )
    if ret_code != 0:
        ret_msg = resp.get("retMsg", "—")
        raise _BybitReportError(f"{chunk_info} retCode={ret_code}, retMsg={ret_msg}")
    result = resp.get("result")
    if not isinstance(result, dict):
        raise _BybitReportError(
            f"{chunk_info} retCode=0, retMsg=нет достоверного result: "
            f"{type(result).__name__}"
        )
    rows = result.get("list")
    if not isinstance(rows, list):
        raise _BybitReportError(
            f"{chunk_info} retCode=0, retMsg=result.list не является списком: "
            f"{type(rows).__name__}"
        )
    return rows


def _next_page_cursor(resp: dict, chunk_info: str) -> str:
    """
    Токен продолжения страницы closed-PnL из ``result["nextPageCursor"]``.

    Вызывается только по уже проверенному _validate_resp ответу, но собственную
    проверку ``result`` не опускает: молча отдать ``""`` по недостоверному ответу
    значило бы объявить выборку законченной без доказательства.

    Возвращает ``""``, когда следующей страницы нет: ключ отсутствует, ``None``
    или пустая строка. Непустой токен возвращается ровно тем значением, которое
    прислала биржа: ни trim, ни смена регистра, ни любая другая нормализация к
    нему не применяются — следующий запрос обязан получить именно его.

    Токен непонятного типа (число, ``bool``, список) окончанием выборки не
    является и подставлять вместо него ``""`` нельзя: это молча обрезало бы
    отчёт. Такой ответ — ошибка.
    """
    result = resp.get("result")
    if not isinstance(result, dict):
        raise _BybitReportError(
            f"{chunk_info} retCode=0, retMsg=нет достоверного result: "
            f"{type(result).__name__}"
        )
    raw = result.get("nextPageCursor")
    if raw is None:
        return ""
    if not isinstance(raw, str):
        raise _BybitReportError(
            f"{chunk_info} retCode=0, retMsg=недоказанный тип nextPageCursor: "
            f"{type(raw).__name__}"
        )
    return raw


async def fetch_closed_pnl_rows(start_ms: int, end_ms: int) -> list:
    """
    Полная выборка строк closed-PnL одного интервала биржи (не более 7 суток).

    Читает страницы по официальному контракту Bybit V5: токен продолжения —
    ``result["nextPageCursor"]``, следующий запрос получает его параметром
    ``cursor``. Окончание выборки доказано только пустым токеном.

    Fail-closed вместо частичного результата, потому что занижённый набор строк
    неотличим от правдивого: аномальный ответ страницы, пустая страница с
    непустым продолжением, повторно выданный токен (иначе цикл шёл бы бесконечно)
    и незавершённая за ``_MAX_PAGES`` страниц пагинация поднимают
    _BybitReportError. Возвращённый список — это доказанно полный интервал.
    """
    chunk_info = f"[{start_ms}–{end_ms}]"
    rows: list = []
    cursor = ""
    seen_cursors: set[str] = set()

    for _ in range(_MAX_PAGES):
        kw: dict = dict(
            category="linear",
            startTime=start_ms,
            endTime=end_ms,
            limit=_PAGE_LIMIT,
        )
        if cursor:
            kw["cursor"] = cursor
        resp = await bybit_call(session.get_closed_pnl, **kw)
        page_rows = _validate_resp(resp, start_ms, end_ms)
        next_cursor = _next_page_cursor(resp, chunk_info)
        rows.extend(page_rows)

        if not next_cursor:
            return rows
        if not page_rows:
            raise _BybitReportError(
                f"{chunk_info} retCode=0, retMsg=пустая страница с непустым "
                "nextPageCursor: окончание выборки не доказано"
            )
        if next_cursor in seen_cursors:
            raise _BybitReportError(
                f"{chunk_info} retCode=0, retMsg=Bybit повторил уже "
                "использованный nextPageCursor: выборка страниц не сходится"
            )
        seen_cursors.add(next_cursor)
        cursor = next_cursor
        await asyncio.sleep(0.1)

    raise _BybitReportError(
        f"{chunk_info} retCode=0, retMsg=пагинация не завершилась за "
        f"{_MAX_PAGES} стр.: интервал получен не полностью"
    )


def _historical_risk_usd(
    trade: dict,
    risk_evidence: dict | None,
    exit_evidence: dict | None = None,
):
    """
    Доказанный исторический риск закрытой сделки в USDT либо ``None``.

    Знаменатель R берётся ТОЛЬКО из durable-записи входа этой самой сделки
    (``ENTRY_PLACED.planned_risk_usdt``) и находится одним из двух точных путей:

    1. прямой — точная пара ``(symbol, orderId)`` closed-PnL строки совпадает с
       ``(symbol, order_id)`` сохранённого входа;
    2. связь выхода — та же точная пара совпадает с
       ``(symbol, exit_order_id)`` события ``EXIT_ORDER_BOUND``, записанного до
       закрытия позиции. Само событие уже несёт риск того входа, который этот
       защитный ордер закрывал.

    Оба пути — точное совпадение идентификаторов, а не поиск похожего. Текущий
    глобальный риск, текущий конфиг, символ сам по себе, время закрытия,
    сторона, цена, объём, источник и близость записей знаменателем не являются:
    они описывают «сейчас» или «похоже», а не ту сделку. Именно использование
    текущего риска делало исторический R зависимым от последующей команды /risk,
    а корреляция по похожести способна приписать сделке чужой риск.

    ``None`` означает «риск этой сделки не доказан», и это правдивый ответ:
    реконструировать его подстановкой любого другого значения запрещено.
    """
    if not isinstance(trade, dict):
        return None
    symbol = normalize_symbol(trade.get("symbol"))
    if not symbol:
        return None
    raw_order_id = trade.get("orderId")
    if not isinstance(raw_order_id, str):
        return None
    order_id = raw_order_id.strip()
    if not order_id:
        return None

    if risk_evidence:
        direct = risk_evidence.get((symbol, order_id))
        if direct is not None:
            return direct

    if not exit_evidence:
        return None
    return exit_evidence.get((symbol, order_id))


def _format_r(value: float) -> str:
    """R с двумя знаками и без хвостовых нулей: ``-4.6R``, ``-6.32R``, ``+4R``.

    Двух знаков достаточно, чтобы результат деления на доказанный риск не
    округлялся до неразличимости, а обрезка хвостовых нулей выполняется только в
    дробной части — целые нули значимы (``+100.00`` обязан остаться ``+100R``).
    """
    whole, _, frac = f"{value:+.2f}".partition(".")
    frac = frac.rstrip("0")
    return f"{whole}.{frac}R" if frac else f"{whole}R"


# ── Нативный XLSX месячного отчёта ───────────────────────────────────────────
# S5 меняет ТОЛЬКО представление/экспорт месячного файла. Источник данных,
# пагинация, границы месяца, порядок сделок, доказанный исторический R и расчёт
# PnL/winrate не меняются — здесь лишь перевод уже посчитанных значений в
# нативную книгу Excel.
_SHEET_TITLE = "Сделки"
_XLSX_HEADERS = (
    "Дата", "Инструмент", "Side", "Entry", "Exit", "PnL USDT", "R", "Источник",
)
# Дата — настоящая datetime-ячейка Excel; формат меняет только отображение.
_DATE_FORMAT = "DD.MM.YYYY HH:MM"
# Цена сохраняет полную точность значения; отображение — динамические знаки.
_PRICE_FORMAT = "0.########"
# PnL остаётся числом; знак и цвет задаются форматом, значение не трогается.
_PNL_FORMAT = "[Green]+0.00;[Red]-0.00;0.00"
_R_FORMAT = "0.00"
# Недоказанный риск в экспорте — операторское тире, а не UNKNOWN/0/0R.
_R_UNKNOWN_DISPLAY = "—"
_COLUMN_WIDTHS = {
    "A": 19, "B": 15, "C": 11, "D": 15, "E": 15, "F": 14, "G": 11, "H": 20,
}
_HEADER_FONT = Font(bold=True)
_HEADER_FILL = PatternFill(
    start_color="FFD9E1F2", end_color="FFD9E1F2", fill_type="solid"
)
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


class _ReportExportError(Exception):
    """Недостоверное значение экспорта: файл не формируется, а не подменяется."""


def _finite_number(value, field: str) -> float:
    """Строгая конвертация authoritative-значения отчёта в конечное число.

    Малформед не превращается в ноль: пустая строка, мусор, ``bool``,
    нефинитное (NaN/Inf) значение или неподдерживаемый тип поднимают
    :class:`_ReportExportError`, и файл целиком не формируется. Фабриковать
    число вместо правды запрещено — оператор должен уметь сортировать,
    фильтровать и суммировать эти колонки без подмены значений.
    """
    if isinstance(value, bool):
        raise _ReportExportError(f"{field}: булево значение не является числом")
    if isinstance(value, (int, float)):
        num = float(value)
    elif isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            raise _ReportExportError(f"{field}: пустое значение не является числом")
        try:
            num = float(stripped)
        except (ValueError, TypeError):
            raise _ReportExportError(
                f"{field}: недостоверное число {value!r}"
            ) from None
    else:
        raise _ReportExportError(
            f"{field}: неподдерживаемый тип {type(value).__name__}"
        )
    if not math.isfinite(num):
        raise _ReportExportError(f"{field}: нефинитное значение {value!r}")
    return num


def _display_source(source) -> str:
    """Операторское отображение источника: снимает только ведущий тег ``#``.

    Технический маркер ``#prodtest`` показывается как ``prodtest``. Встроенные
    ``#`` внутри остального текста не трогаются, а сама персистентная запись
    источника не меняется — это только отображаемое значение экспорта.
    """
    text = "" if source is None else str(source)
    if text.startswith("#"):
        text = text[1:]
    return text


def _write_text_cell(cell, text):
    """Пишет литеральную строку, не давая ей стать формулой Excel.

    Значение, начинающееся с ``=``, openpyxl иначе пометил бы как формулу.
    Принудительный строковый тип держит внешний текст (``=1+1``, ``+``, ``-``,
    ``@``) обычной строкой при открытии книги в Excel.
    """
    cell.value = "" if text is None else str(text)
    cell.data_type = "s"
    return cell


def _build_report_workbook(rows: list) -> io.BytesIO:
    """Строит месячный ``.xlsx`` в памяти и возвращает его как BytesIO.

    Одна closed-PnL строка — одна строка листа; агрегации здесь нет. ``Дата`` —
    настоящая datetime-ячейка, ``Entry``/``Exit``/``PnL USDT`` и доказанный R —
    числовые ячейки, недоказанный R — тире. Строгая числовая конвертация
    поднимает :class:`_ReportExportError` наружу как ошибку отчёта: частичный
    или битый файл оператору не отправляется.
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = _SHEET_TITLE

    worksheet.append(list(_XLSX_HEADERS))
    for cell in worksheet[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _HEADER_ALIGN

    for record in rows:
        idx = worksheet.max_row + 1

        date_cell = worksheet.cell(row=idx, column=1, value=record["dt"])
        date_cell.number_format = _DATE_FORMAT

        _write_text_cell(worksheet.cell(row=idx, column=2), record["symbol"])
        _write_text_cell(worksheet.cell(row=idx, column=3), record["side"])

        entry_cell = worksheet.cell(
            row=idx, column=4,
            value=_finite_number(record["entry"], "avgEntryPrice"),
        )
        entry_cell.number_format = _PRICE_FORMAT

        exit_cell = worksheet.cell(
            row=idx, column=5,
            value=_finite_number(record["exit"], "avgExitPrice"),
        )
        exit_cell.number_format = _PRICE_FORMAT

        pnl_cell = worksheet.cell(
            row=idx, column=6, value=_finite_number(record["pnl"], "closedPnl"),
        )
        pnl_cell.number_format = _PNL_FORMAT

        if record["r"] is None:
            _write_text_cell(
                worksheet.cell(row=idx, column=7), _R_UNKNOWN_DISPLAY
            )
        else:
            r_cell = worksheet.cell(
                row=idx, column=7, value=_finite_number(record["r"], "R"),
            )
            r_cell.number_format = _R_FORMAT

        _write_text_cell(
            worksheet.cell(row=idx, column=8), _display_source(record["source"])
        )

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = f"A1:H{worksheet.max_row}"
    for column, width in _COLUMN_WIDTHS.items():
        worksheet.column_dimensions[column].width = width

    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


async def send_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """
    Команда /report [мм.гггг] — отчёт о закрытых сделках за месяц.

    Без аргументов: показывает текстовый список последних 15 сделок.
    С аргументом даты (например, /report 01.2026): отправляет нативную книгу
    Excel (.xlsx) с полной выборкой. Данные получаются чанками по 7 дней для
    обхода лимитов API.
    """
    if str(update.effective_user.id) != ALLOWED_ID: return

    now = datetime.now(timezone.utc)
    now_ms = int(now.timestamp() * 1000)

    if context.args:
        try:
            month_str, year_str = context.args[0].split('.')
            target_date = datetime(int(year_str), int(month_str), 1, tzinfo=timezone.utc)
        except (ValueError, TypeError):
            await update.message.reply_text(
                f"{format_header('⚠️', 'WARNING')}\n\n"
                f"⚠️ <b>Предупреждения</b>\n"
                f"• Неверный формат месяца.\n\n"
                f"{format_action('используйте /report 01.2026')}",
                parse_mode='HTML',
            )
            return
    else:
        target_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    start_ts = int(target_date.replace(day=1, hour=0, minute=0, second=0, microsecond=0).timestamp() * 1000)

    next_month = target_date.replace(day=28) + timedelta(days=4)
    end_date = next_month - timedelta(days=next_month.day)
    end_ts = int(end_date.replace(hour=23, minute=59, second=59, microsecond=0).timestamp() * 1000)
    # Не запрашиваем будущее — зажимаем верхнюю границу текущим моментом
    end_ts = min(end_ts, now_ms)

    month_name = target_date.strftime("%B %Y")
    status_msg = await update.message.reply_text(
        f"{format_header('⏳', 'REPORT')}\n\n"
        f"Собираю данные за {h(month_name)}.",
        parse_mode='HTML',
    )

    status_deleted = False
    current_start = start_ts
    current_end = start_ts   # инициализируем до цикла — доступно в except
    try:
        all_trades = []

        while current_start < end_ts:
            current_end = min(current_start + _CHUNK_MS, end_ts)
            # Полная выборка чанка или ошибка: частичные страницы агрегатом
            # отчёта не становятся.
            all_trades.extend(await fetch_closed_pnl_rows(current_start, current_end))
            current_start = current_end + 1          # шаг на 1 мс — без пробелов и перекрытий
            await asyncio.sleep(0.1)

        if not all_trades:
            await status_msg.edit_text(
                f"{format_header('📊', 'REPORT')}\n\n"
                f"ℹ️ За {h(month_name)} закрытых сделок нет.",
                parse_mode='HTML',
            )
            return

        total_pnl = 0
        wins = 0
        losses = 0
        xlsx_rows = []
        report_lines = []
        # Доказанный риск конкретных входов бота. Читается один раз за отчёт;
        # запись в журнал не производится — backfill историческим риском запрещён.
        risk_evidence = await asyncio.to_thread(get_entry_risk_evidence)
        # Связи защитных ордеров выхода с риском их входа. Записаны наблюдателем
        # ДО закрытия позиции, пока ордер был виден в открытых: только так
        # дочерний SL/TP вообще сохраняет связь со своим входом. Никаких запросов
        # к бирже здесь нет — история ордеров для этого бесполезна.
        exit_evidence = await asyncio.to_thread(get_exit_order_risk_evidence)
        # Аккумуляторы R считаются ТОЛЬКО по сделкам с доказанным риском.
        total_r = 0.0
        r_known = 0

        all_trades.sort(key=lambda x: int(x['updatedTime']), reverse=True)

        for t in all_trades:
            symbol = t['symbol']
            # Authoritative closedPnl проверяется СТРОГО из сырого значения биржи
            # ДО любых производных (total_pnl, wins/losses, R и строка книги).
            # safe_float здесь запрещён: он молча свёл бы missing/None/""/пробелы/
            # мусор/[]/{}/False к 0.0, а True — к 1.0, сфабриковав нулевой PnL и
            # фальшивый R (0 или 0.5). Битое authoritative-значение обязано
            # провалить отчёт целиком (_ReportExportError → общий except), а не
            # подмениться нулём. Легальный ноль (0/"0"/"0.00") остаётся числом.
            pnl = _finite_number(t.get('closedPnl'), 'closedPnl')
            ts = int(t.get('updatedTime', 0))
            short_date = datetime.fromtimestamp(ts / 1000).strftime("%d.%m")

            total_pnl += pnl
            if pnl > 0:
                wins += 1
            elif pnl < 0:
                losses += 1

            trade_risk = _historical_risk_usd(t, risk_evidence, exit_evidence)
            if trade_risk is None:
                # Риск этой сделки не доказан: R недоступен. Ни ноль, ни текущий
                # глобальный риск подстановкой быть не могут.
                r_text = UNKNOWN
                r_value = None
            else:
                r_val = pnl / trade_risk
                total_r += r_val
                r_known += 1
                r_text = _format_r(r_val)
                r_value = r_val
            src = get_source_at_time(symbol, ts)

            # Одна closed-PnL строка = одна строка книги. Дата — настоящий
            # datetime того же базиса, что и текст отчёта; Entry/Exit остаются
            # сырыми значениями биржи и строго конвертируются только при сборке
            # книги (путь файла), чтобы текстовый отчёт без аргумента не менялся.
            xlsx_rows.append({
                "dt": datetime.fromtimestamp(ts / 1000),
                "symbol": symbol,
                "side": t['side'],
                "entry": t['avgEntryPrice'],
                "exit": t['avgExitPrice'],
                "pnl": pnl,
                "r": r_value,
                "source": src,
            })

            icon = "🟢" if pnl >= 0 else "🔴"
            line = (
                f"{icon} {h(short_date)} · {h(symbol)} · "
                f"{pnl:+.1f} USDT · {h(r_text)} · {h(src)}"
            )
            report_lines.append(line)

        total_trades = wins + losses
        winrate = (wins / total_trades * 100) if total_trades > 0 else 0
        # Агрегат R правдиво сообщает свою полноту: без доказанных сделок он не
        # выводится вовсе, при частичном покрытии рядом стоит охват.
        if r_known == 0:
            total_r_text = UNKNOWN
        elif r_known == len(all_trades):
            total_r_text = f"{total_r:+.2f}R"
        else:
            total_r_text = (
                f"{total_r:+.2f}R (по {r_known} из {len(all_trades)} сделок)"
            )

        cmd_example = f"/report {target_date.strftime('%m.%Y')}"
        summary_block = format_value_block([
            ("PnL", f"{total_pnl:+.2f} USDT"),
            ("R", total_r_text),
            ("Winrate", f"{winrate:.1f}% ({wins}W / {losses}L)"),
            ("Сделки", total_trades),
        ])

        header = (
            f"{format_header('📊', 'REPORT')}\n"
            f"Период: {h(month_name)}\n\n"
            f"📊 <b>Итоги</b>\n"
            f"{summary_block}\n\n"
            f"📅 Другой месяц: <code>{h(cmd_example)}</code>"
        )

        status_deleted = True
        await status_msg.delete()

        if context.args:
            # Нативная книга Excel строится в памяти. Ошибка строгой числовой
            # конвертации поднимется наружу в общий except и станет ошибкой
            # отчёта — битый/частичный файл оператору не уходит.
            workbook = _build_report_workbook(xlsx_rows)
            await update.message.reply_document(
                document=workbook,
                filename=f"Report_{target_date.strftime('%m_%Y')}.xlsx",
                caption=header,
                parse_mode='HTML'
            )
        else:
            short_list = "\n".join(report_lines[:15])
            await update.message.reply_text(
                f"{header}\n\n📋 <b>Последние 15</b>\n{short_list}",
                parse_mode='HTML',
            )

    except Exception as e:
        logging.exception(
            "Report error for %s (chunk %s–%s): %s",
            month_name, current_start, current_end, e,
        )
        err_text = format_error_message(
            "Не удалось сформировать отчёт Bybit.",
            context=month_name,
            detail=format_bybit_error_detail(e),
            action="повторите запрос отчёта позже",
        )
        if not status_deleted:
            try:
                await status_msg.edit_text(err_text, parse_mode='HTML')
                return
            except Exception:
                pass
        await update.message.reply_text(err_text, parse_mode='HTML')
