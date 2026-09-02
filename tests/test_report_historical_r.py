"""
LIVE-FIX2/LIVE-FIX4 — исторический R в /report считается только по доказанному
риску сделки и только по точной идентичности ордеров.

Проверяемые сценарии (LIVE-FIX2):
1. Доказанный риск конкретного входа даёт R, не зависящий от текущего /risk.
2. Один и тот же отчёт при глобальном риске 1/5/10/50 совпадает байт в байт.
3. Сделка без доказанного риска получает UNKNOWN, а не R по текущему риску.
4. Агрегат R собирается только из доказанных сделок и правдиво сообщает охват.
5. PnL, winrate и число сделок не зависят от доказанности риска.
6. XLSX-экспорт использует ту же семантику: доказанный R числом, недоказанный —
   тире.
7. Совпадение только по символу доказательством не является.
8. Карта доказанного риска строится строго по паре (symbol, order_id).
9. Риск, конечный для Decimal, но дающий inf/0.0 во float, доказательством не
   является: R остаётся UNKNOWN, а не превращается в фальшивый 0R.

Проверяемые сценарии (LIVE-FIX4 — связь защитного ордера выхода):
10. closed-PnL (symbol, orderId) совпадает с (symbol, exit_order_id) durable
    события EXIT_ORDER_BOUND, записанного ДО закрытия → доказанный риск входа.
11. Закрывающий orderId не равен входному, и это нормально: связь сохранена
    заранее, потому что после исполнения SL/TP биржа родства не отдаёт.
12. Прямой путь входа (LIVE-FIX2) остаётся приоритетным; связь выхода его не
    перехватывает даже при противоречивом значении.
13. Чужой символ, чужой exit orderId и отсутствие связи дают UNKNOWN, а не R по
    текущему глобальному риску.
14. Противоречивый риск по одному (symbol, exit_order_id) fail-closed.
15. Старые сделки без связи остаются UNKNOWN: backfill запрещён.
16. Месячный scan истории ордеров из runtime отчёта удалён вместе с мёртвым
    ancestry-кодом: у SL/TP-детей orderLinkId и parentOrderLinkId пустые.

Все зависимости Bybit/Telegram замокированы; сетевых вызовов нет.
"""

import io
import math
import sys
import os
from datetime import datetime, timezone
from decimal import Decimal
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl

# ── Mock heavy deps before any project import ────────────────────────────────
for _mod in [
    "telegram", "telegram.ext", "telegram.request", "telegram.error",
    "pybit", "pybit.unified_trading",
    "dotenv", "colorama",
]:
    sys.modules.setdefault(_mod, MagicMock())

_UID = "123"

_cfg = MagicMock()
_cfg.ALLOWED_ID = _UID
_cfg.DATA_DIR = Path(__file__).resolve().parent.parent / "data"
sys.modules.setdefault("core.config", _cfg)

for _mod in ["core.trading_core", "core.bybit_call", "core.database", "handlers.orders"]:
    sys.modules.setdefault(_mod, MagicMock())

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pytest  # noqa: E402

import handlers.reporting as reporting  # noqa: E402
from core.journal import (  # noqa: E402
    ENTRY_PLACED, EXIT_BINDING_SOURCE_OPEN_ORDERS, EXIT_KIND_TP,
    EXIT_ORDER_BOUND, UNKNOWN, _proven_risk_usdt, get_entry_risk_evidence,
    get_exit_order_risk_evidence,
)


# ── Фиксированное «сейчас»: месяц отчёта не зависит от даты запуска ──────────

class _FixedDatetime(datetime):
    """datetime с детерминированным now(): отчёт всегда за февраль 2026."""

    @classmethod
    def now(cls, tz=None):
        return datetime(2026, 2, 20, 12, 0, 0, tzinfo=timezone.utc)


# ── Helpers ─────────────────────────────────────────────────────────────────

def _entry(symbol="BTCUSDT", order_id="OID-1", risk=1.0, **extra):
    """Durable-запись входа бота с запланированным риском."""
    ev = {
        "event": ENTRY_PLACED, "symbol": symbol, "side": "LONG",
        "order_id": order_id, "planned_risk_usdt": risk, "ts": 1000.0,
    }
    ev.update(extra)
    return ev

def _trade(symbol="BTCUSDT", pnl="-4.6", order_id="OID-1", ts=1770000000000):
    """Строка закрытой сделки в форме ответа get_closed_pnl."""
    return {
        "symbol": symbol, "closedPnl": pnl, "updatedTime": str(ts),
        "side": "Sell", "avgEntryPrice": "100", "avgExitPrice": "95",
        "orderId": order_id,
    }


def _page(trades):
    """Одна страница ответа Bybit без продолжения.

    Пагинация closed-PnL в этом stage не меняется: известный дефект выбора токена
    продолжения отнесён к LIVE-FIX5 и здесь намеренно не проверяется.
    """
    return {"retCode": 0, "retMsg": "OK", "result": {"list": trades}}


def _bound(symbol="ETHUSDT", exit_order_id="CLOSE-TP-1", risk=3.0,
           exit_kind=EXIT_KIND_TP, entry_order_id="ENTRY-1", **extra):
    """Durable-событие связи защитного ордера выхода с риском своего входа."""
    ev = {
        "event": EXIT_ORDER_BOUND, "symbol": symbol, "side": "Buy",
        "position_idx": 0, "entry_order_id": entry_order_id,
        "exit_order_id": exit_order_id, "exit_kind": exit_kind,
        "planned_risk_usdt": risk, "trigger_price": "1873.5",
        "binding_source": EXIT_BINDING_SOURCE_OPEN_ORDERS, "ts": 1000.0,
    }
    ev.update(extra)
    return ev


def _summary_value(text, label):
    """Значение строки блока итогов: отступы и закрывающий тег не важны."""
    prefix = f"{label}:"
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith(prefix):
            return stripped[len(prefix):].removesuffix("</code>").strip()
    raise AssertionError(f"в итогах нет строки {label}: {text}")


def _summary_r(text):
    """Значение строки R из блока итогов."""
    return _summary_value(text, "R")


def _summary_trades(text):
    """Значение строки «Сделки» из блока итогов."""
    return _summary_value(text, "Сделки")


async def _run_report(trades, evidence, *, args=None, exit_evidence=None,
                      calls=None):
    """
    Выполняет send_report на замокированных Bybit/Telegram.

    ``evidence`` — карта прямого пути LIVE-FIX2 ``(symbol, order_id) → риск``,
    ``exit_evidence`` — карта связи выхода LIVE-FIX4
    ``(symbol, exit_order_id) → риск``. Разрешение R выполняется настоящим кодом
    отчёта: подменяются только читатели журнала, а не сама логика приоритета.

    В ``calls``, если он передан, складываются ``(fn, kwargs)`` каждого обращения
    к Bybit — так проверяется, что отчёт читает только closed-PnL.

    Возвращает (text, workbook): текст последнего сообщения и загруженную
    openpyxl книгу месячного экспорта (``None``, если документ не отправлялся).
    """
    status_msg = MagicMock()
    status_msg.edit_text = AsyncMock()
    status_msg.delete = AsyncMock()

    update = MagicMock()
    update.effective_user.id = _UID
    update.message.reply_text = AsyncMock(return_value=status_msg)
    update.message.reply_document = AsyncMock()

    context = MagicMock()
    context.args = list(args or [])

    # Сделки возвращаются один раз: чанки за месяц не должны их дублировать.
    pages = [_page(list(trades))]

    async def _fake_call(fn, **kw):
        if calls is not None:
            calls.append((fn, dict(kw)))
        return pages.pop(0) if pages else _page([])

    with patch.object(reporting, "ALLOWED_ID", _UID), \
            patch.object(reporting, "datetime", _FixedDatetime), \
            patch.object(reporting, "bybit_call", new=AsyncMock(side_effect=_fake_call)), \
            patch.object(reporting, "get_source_at_time", return_value="TG"), \
            patch.object(reporting, "get_entry_risk_evidence", return_value=dict(evidence)), \
            patch.object(reporting, "get_exit_order_risk_evidence",
                         return_value=dict(exit_evidence or {})), \
            patch("asyncio.sleep", new=AsyncMock()):
        await reporting.send_report(update, context)

    # Ошибочный путь отчёта не должен срабатывать ни в одном сценарии.
    assert status_msg.delete.await_count == 1, (
        f"отчёт не дошёл до отправки: {status_msg.edit_text.call_args_list}"
    )

    text = ""
    if update.message.reply_text.await_count > 1:
        text = update.message.reply_text.call_args_list[-1].args[0]

    workbook = None
    if update.message.reply_document.await_count:
        kwargs = update.message.reply_document.call_args.kwargs
        document = kwargs["document"]
        assert isinstance(document, io.BytesIO)
        document.seek(0)
        workbook = openpyxl.load_workbook(document)
        text = kwargs["caption"]

    return text, workbook


# ── Доказанный риск сделки → правдивый R ────────────────────────────────────

class TestProvenHistoricalR:

    @pytest.mark.asyncio
    async def test_risk_1_gives_full_r(self):
        """pnl=-4.6 при доказанном риске 1 USDT → -4.6R."""
        text, _ = await _run_report(
            [_trade(pnl="-4.6", order_id="OID-1")],
            {("BTCUSDT", "OID-1"): 1.0},
        )
        assert "-4.6R" in text
        assert UNKNOWN not in text

    @pytest.mark.asyncio
    @pytest.mark.parametrize("current_risk", [1.0, 5.0, 10.0, 50.0])
    async def test_r_does_not_follow_current_global_risk(self, current_risk):
        """Текущий /risk 1/5/10/50 не меняет отчёт по доказанной сделке."""
        with patch("core.database.get_global_risk", return_value=current_risk):
            text, _ = await _run_report(
                [_trade(pnl="-4.6", order_id="OID-1")],
                {("BTCUSDT", "OID-1"): 1.0},
            )
        assert "-4.6R" in text
        # Итог считается по тому же доказанному риску, а не по текущему.
        assert _summary_r(text) == "-4.60R"

    @pytest.mark.asyncio
    async def test_risk_5_divides_historical_pnl(self):
        """pnl=-31.6 при доказанном риске 5 USDT → -6.32R."""
        text, _ = await _run_report(
            [_trade(symbol="GRVTUSDT", pnl="-31.6", order_id="OID-G")],
            {("GRVTUSDT", "OID-G"): 5.0},
        )
        assert "-6.32R" in text

    def test_current_global_risk_is_not_imported(self):
        """get_global_risk больше не участвует в отчёте — регрессия дефекта."""
        assert not hasattr(reporting, "get_global_risk")

    def test_format_r_trims_only_fraction(self):
        """Хвостовые нули режутся лишь в дробной части: целые нули значимы."""
        assert reporting._format_r(-4.6) == "-4.6R"
        assert reporting._format_r(-6.32) == "-6.32R"
        assert reporting._format_r(4.0) == "+4R"
        assert reporting._format_r(100.0) == "+100R"


# ── Недоказанный риск → UNKNOWN, без реконструкции ──────────────────────────

class TestUnknownHistoricalR:

    @pytest.mark.asyncio
    async def test_legacy_trade_without_evidence_is_unknown(self):
        """Legacy-сделка без доказанного риска не получает R по текущему риску."""
        with patch("core.database.get_global_risk", return_value=5.0):
            text, _ = await _run_report([_trade(pnl="-4.6", order_id="OID-OLD")], {})
        assert UNKNOWN in text
        assert "-4.6R" not in text
        assert "-0.92R" not in text          # деление на текущий риск запрещено
        assert "-4.6 USDT" in text           # PnL остаётся доступен

    @pytest.mark.asyncio
    async def test_symbol_match_alone_is_not_evidence(self):
        """Совпал только символ, orderId другой → UNKNOWN."""
        text, _ = await _run_report(
            [_trade(symbol="BTCUSDT", pnl="-4.6", order_id="CLOSE-ORDER")],
            {("BTCUSDT", "OID-1"): 1.0},
        )
        assert UNKNOWN in text
        assert "-4.6R" not in text

    @pytest.mark.asyncio
    async def test_trade_without_order_id_is_unknown(self):
        """Строка без orderId сопоставлению не подлежит → UNKNOWN."""
        trade = _trade(pnl="-4.6")
        del trade["orderId"]
        text, _ = await _run_report(trade and [trade], {("BTCUSDT", "OID-1"): 1.0})
        assert UNKNOWN in text

    @pytest.mark.asyncio
    async def test_aggregate_unknown_when_nothing_proven(self):
        """Ни одной доказанной сделки → агрегат R не выводится как число."""
        text, _ = await _run_report(
            [_trade(pnl="-4.6", order_id="A"), _trade(pnl="-31.6", order_id="B")],
            {},
        )
        assert _summary_r(text) == UNKNOWN

    @pytest.mark.asyncio
    async def test_overflow_risk_trade_is_unknown_not_zero_r(self):
        """Риск «1e9999» проходил Decimal-проверку и давал фальшивый 0R."""
        # Evidence строится настоящим journal-хелпером: подставлять сюда готовое
        # число значило бы обойти ровно ту проверку, которую чинит эта правка.
        evidence = get_entry_risk_evidence([_entry(order_id="OID-1", risk="1e9999")])
        text, _ = await _run_report([_trade(pnl="-4.6", order_id="OID-1")], evidence)
        assert f"-4.6 USDT · {UNKNOWN}" in text
        assert "0R" not in text                    # pnl / inf → +0.0R запрещён

    @pytest.mark.asyncio
    async def test_overflow_risk_trade_keeps_pnl_winrate_and_count(self):
        """Сделка с недоказанным риском остаётся в PnL, winrate и счёте сделок."""
        evidence = get_entry_risk_evidence([_entry(order_id="OID-1", risk="1e9999")])
        text, _ = await _run_report([_trade(pnl="-4.6", order_id="OID-1")], evidence)
        assert "-4.60 USDT" in text
        assert "0.0% (0W / 1L)" in text
        assert _summary_trades(text) == "1"

    @pytest.mark.asyncio
    async def test_aggregate_unknown_when_all_risk_overflows(self):
        """Месяц, где всё evidence — overflow/underflow: агрегат UNKNOWN, не 0R."""
        evidence = get_entry_risk_evidence([
            _entry(order_id="A", risk="1e9999"),
            _entry(order_id="B", risk="1e-9999"),
        ])
        assert evidence == {}
        text, _ = await _run_report(
            [
                _trade(pnl="-4.6", order_id="A", ts=1770000000000),
                _trade(pnl="-31.6", order_id="B", ts=1770000100000),
            ],
            evidence,
        )
        assert _summary_r(text) == UNKNOWN
        assert "0.00R" not in text

    @pytest.mark.asyncio
    async def test_xlsx_overflow_risk_is_em_dash_not_zero(self):
        """XLSX для overflow-риска: R — тире, а не 0; PnL сохранён числом."""
        evidence = get_entry_risk_evidence([_entry(order_id="OID-1", risk="1e9999")])
        _, workbook = await _run_report(
            [_trade(pnl="-4.6", order_id="OID-1")], evidence, args=["02.2026"],
        )
        ws = workbook.active
        assert ws.cell(row=2, column=7).value == "—"       # R не доказан
        assert ws.cell(row=2, column=7).value != UNKNOWN
        assert ws.cell(row=2, column=6).value == -4.6       # PnL сохранён числом
        assert ws.cell(row=2, column=6).data_type == "n"


# ── Агрегаты: R по доказанным, PnL/winrate/счёт по всем ─────────────────────

class TestAggregates:

    @pytest.mark.asyncio
    async def test_partial_coverage_is_stated_truthfully(self):
        """Один доказанный из двух: агрегат R указывает охват, PnL полный."""
        text, _ = await _run_report(
            [
                _trade(symbol="BTCUSDT", pnl="-4.6", order_id="OID-1", ts=1770000000000),
                _trade(symbol="GRVTUSDT", pnl="-31.6", order_id="OID-X", ts=1770000100000),
            ],
            {("BTCUSDT", "OID-1"): 1.0},
        )
        assert "-4.60R (по 1 из 2 сделок)" in text
        assert "-36.20 USDT" in text                 # PnL по всем сделкам
        assert "0.0% (0W / 2L)" in text              # winrate по всем сделкам
        assert _summary_trades(text) == "2"

    @pytest.mark.asyncio
    async def test_full_coverage_has_no_coverage_note(self):
        """Все сделки доказаны → в агрегате нет пометки охвата."""
        text, _ = await _run_report(
            [
                _trade(symbol="BTCUSDT", pnl="-4.6", order_id="OID-1", ts=1770000000000),
                _trade(symbol="GRVTUSDT", pnl="-31.6", order_id="OID-G", ts=1770000100000),
            ],
            {("BTCUSDT", "OID-1"): 1.0, ("GRVTUSDT", "OID-G"): 5.0},
        )
        assert "-10.92R" in text                     # -4.6R + -6.32R
        assert "из 2 сделок" not in text


# ── XLSX: та же семантика ────────────────────────────────────────────────────

class TestXlsxOutput:

    @pytest.mark.asyncio
    async def test_xlsx_mixes_proven_number_and_em_dash(self):
        """В XLSX доказанный R — число, недоказанный — тире, а не 0."""
        _, workbook = await _run_report(
            [
                _trade(symbol="BTCUSDT", pnl="-4.6", order_id="OID-1", ts=1770000000000),
                _trade(symbol="GRVTUSDT", pnl="-31.6", order_id="OID-X", ts=1770000100000),
            ],
            {("BTCUSDT", "OID-1"): 1.0},
            args=["02.2026"],
        )
        ws = workbook.active
        assert [c.value for c in ws[1]] == [
            "Дата", "Инструмент", "Side", "Entry", "Exit", "PnL USDT", "R", "Источник",
        ]
        rows = {ws.cell(row=r, column=2).value: r for r in range(2, ws.max_row + 1)}
        grvt, btc = rows["GRVTUSDT"], rows["BTCUSDT"]
        # GRVT: риск не доказан → тире, PnL сохранён числом.
        assert ws.cell(row=grvt, column=7).value == "—"
        assert ws.cell(row=grvt, column=6).value == -31.6
        # BTC: риск доказан → R числом.
        assert ws.cell(row=btc, column=7).value == -4.6
        assert ws.cell(row=btc, column=7).data_type == "n"


# ── Карта доказанного риска в журнале ───────────────────────────────────────

class TestEntryRiskEvidence:

    def test_proven_entry_becomes_evidence(self):
        """ENTRY_PLACED с символом, orderId и риском → пара (symbol, order_id)."""
        got = get_entry_risk_evidence([_entry(order_id="OID-1", risk=2.5)])
        assert got == {("BTCUSDT", "OID-1"): 2.5}

    def test_symbol_is_normalized(self):
        """Символ приводится к верхнему регистру без пробелов."""
        got = get_entry_risk_evidence([_entry(symbol=" btcusdt ")])
        assert ("BTCUSDT", "OID-1") in got

    @pytest.mark.parametrize("risk", [0, 0.0, -1, None, "", "—", True, False, "abc",
                                      float("nan"), float("inf"), "1e9999",
                                      "-1e9999", "1e-9999", Decimal("1e9999"),
                                      Decimal("1e-9999")])
    def test_unproven_risk_creates_no_evidence(self, risk):
        """Ноль, отрицательный, bool, NaN, Infinity, overflow и мусор — не риск."""
        assert get_entry_risk_evidence([_entry(risk=risk)]) == {}

    @pytest.mark.parametrize("raw", ["1e9999", "-1e9999", "9" * 400])
    def test_decimal_finite_but_float_overflow_is_unproven(self, raw):
        """Конечный для Decimal риск, дающий inf во float, доказательством не является."""
        # Именно эта комбинация и создавала дефект: Decimal-проверка проходила...
        assert Decimal(raw).is_finite()
        assert math.isinf(float(Decimal(raw)))
        # ...а знаменателем становился inf, из которого /report получал 0R.
        assert _proven_risk_usdt(raw) is None

    @pytest.mark.parametrize("raw", ["1e-9999", "1e-400"])
    def test_decimal_underflow_to_zero_is_unproven(self, raw):
        """Риск, схлопывающийся во float в 0.0, знаменателем быть не может."""
        assert float(Decimal(raw)) == 0.0
        assert _proven_risk_usdt(raw) is None

    @pytest.mark.parametrize("raw", ["0.5", "1", "5", "31.6", "1e-6", "1e300"])
    def test_finite_positive_values_survive(self, raw):
        """Обычные конечные положительные значения регрессии не получили."""
        got = _proven_risk_usdt(raw)
        assert got is not None
        assert math.isfinite(got)
        assert got > 0

    def test_missing_risk_key_creates_no_evidence(self):
        """Старое событие без planned_risk_usdt evidence не даёт."""
        ev = _entry()
        del ev["planned_risk_usdt"]
        assert get_entry_risk_evidence([ev]) == {}

    @pytest.mark.parametrize("order_id", ["", "   ", None, 12345])
    def test_missing_order_id_creates_no_evidence(self, order_id):
        """Без точного строкового orderId сопоставление невозможно."""
        ev = _entry()
        ev["order_id"] = order_id
        assert get_entry_risk_evidence([ev]) == {}

    def test_other_events_are_ignored(self):
        """Не-ENTRY_PLACED события риском входа не являются."""
        ev = _entry()
        ev["event"] = "RECONCILED"
        assert get_entry_risk_evidence([ev]) == {}

    def test_string_risk_is_parsed_exactly(self):
        """Риск строкой разбирается через Decimal без потери точности."""
        assert _proven_risk_usdt("0.5") == 0.5
        assert _proven_risk_usdt(" 5 ") == 5.0


# ── LIVE-FIX4: связь защитного ордера выхода с риском входа ─────────────────

# Доказанный вход production-сценария: ETH, риск 3 USDT.
# Связь записана наблюдателем ДО закрытия, пока TP был виден в открытых ордерах.
_ETH_EXIT_RISK = {("ETHUSDT", "CLOSE-TP-1"): 3.0}
# Тот же вход по прямому пути: closed-PnL строка ссылается на ордер входа.
_ETH_DIRECT_RISK = {("ETHUSDT", "ENTRY-1"): 3.0}


class TestExitBindingHistoricalR:

    @pytest.mark.asyncio
    async def test_child_close_order_resolves_r_via_exit_binding(self):
        """closedPnl=-0.12 при связанном риске 3 USDT → -0.04R."""
        text, _ = await _run_report(
            [_trade(symbol="ETHUSDT", pnl="-0.12", order_id="CLOSE-TP-1")],
            {},                                   # прямой путь не совпадает
            exit_evidence=_ETH_EXIT_RISK,
        )
        assert "-0.04R" in text
        assert UNKNOWN not in text

    @pytest.mark.asyncio
    async def test_close_order_id_differs_from_entry_order_id(self):
        """Закрывающий orderId не равен входному — связь всё равно доказана."""
        text, _ = await _run_report(
            [_trade(symbol="ETHUSDT", pnl="-3", order_id="CLOSE-TP-1")],
            _ETH_DIRECT_RISK,                     # (ETHUSDT, ENTRY-1) — не совпадёт
            exit_evidence=_ETH_EXIT_RISK,
        )
        assert "-1R" in text

    @pytest.mark.asyncio
    @pytest.mark.parametrize("current_risk", [1.0, 3.0, 5.0, 50.0])
    async def test_exit_r_does_not_follow_current_global_risk(self, current_risk):
        """Текущий /risk 1/3/5/50 не меняет R, доказанный связью выхода."""
        with patch("core.database.get_global_risk", return_value=current_risk):
            text, _ = await _run_report(
                [_trade(symbol="ETHUSDT", pnl="-0.12", order_id="CLOSE-TP-1")],
                {},
                exit_evidence=_ETH_EXIT_RISK,
            )
        assert "-0.04R" in text
        assert _summary_r(text) == "-0.04R"

    @pytest.mark.asyncio
    async def test_stop_loss_child_order_resolves_r(self):
        """Закрытие по SL доказывается той же связью, что и по TP: риск входа один."""
        text, _ = await _run_report(
            [_trade(symbol="ETHUSDT", pnl="-3", order_id="CLOSE-SL-1")],
            {},
            exit_evidence={("ETHUSDT", "CLOSE-SL-1"): 3.0},
        )
        assert "-1R" in text

    @pytest.mark.asyncio
    async def test_direct_entry_path_wins_over_exit_binding(self):
        """Прямой точный путь входа остаётся первым и не заменяется связью выхода."""
        text, _ = await _run_report(
            [_trade(symbol="ETHUSDT", pnl="-3", order_id="ENTRY-1")],
            _ETH_DIRECT_RISK,
            # Противоречивая связь по тому же orderId приоритет не перехватывает.
            exit_evidence={("ETHUSDT", "ENTRY-1"): 12.0},
        )
        assert "-1R" in text
        assert "-0.25R" not in text

    @pytest.mark.asyncio
    async def test_xlsx_uses_the_same_resolved_r(self):
        """XLSX использует тот же доказанный R, что и сообщение."""
        _, workbook = await _run_report(
            [_trade(symbol="ETHUSDT", pnl="-3", order_id="CLOSE-TP-1")],
            {},
            exit_evidence=_ETH_EXIT_RISK,
            args=["02.2026"],
        )
        ws = workbook.active
        assert ws.cell(row=2, column=6).value == -3.0    # PnL
        assert ws.cell(row=2, column=7).value == -1.0    # R = -3 / 3
        assert ws.cell(row=2, column=7).data_type == "n"

    @pytest.mark.asyncio
    async def test_partial_coverage_is_stated_truthfully(self):
        """Доказана одна из двух сделок: охват в агрегате правдив, PnL полный."""
        text, _ = await _run_report(
            [
                _trade(symbol="ETHUSDT", pnl="-3", order_id="CLOSE-TP-1",
                       ts=1770000000000),
                _trade(symbol="BTCUSDT", pnl="-4.6", order_id="CLOSE-2",
                       ts=1770000100000),
            ],
            {},
            exit_evidence=_ETH_EXIT_RISK,
        )
        assert "-1.00R (по 1 из 2 сделок)" in text
        assert "-7.60 USDT" in text
        assert _summary_trades(text) == "2"

    @pytest.mark.asyncio
    async def test_pnl_winrate_and_count_do_not_depend_on_binding(self):
        """Связь влияет только на R: PnL, winrate и число сделок те же."""
        trades = [
            _trade(symbol="ETHUSDT", pnl="-3", order_id="CLOSE-TP-1",
                   ts=1770000000000),
            _trade(symbol="BTCUSDT", pnl="6", order_id="CLOSE-2",
                   ts=1770000100000),
        ]
        bound, _ = await _run_report(trades, {}, exit_evidence=_ETH_EXIT_RISK)
        unbound, _ = await _run_report(trades, {}, exit_evidence={})
        for text in (bound, unbound):
            assert "+3.00 USDT" in text
            assert "50.0% (1W / 1L)" in text
            assert _summary_trades(text) == "2"


class TestExitBindingUnproven:

    @pytest.mark.asyncio
    async def test_no_binding_is_unknown_without_current_risk_fallback(self):
        """Связи нет → UNKNOWN, а не R по текущему глобальному риску."""
        with patch("core.database.get_global_risk", return_value=3.0):
            text, _ = await _run_report(
                [_trade(symbol="ETHUSDT", pnl="-0.12", order_id="CLOSE-TP-1")],
                {},
                exit_evidence={},
            )
        assert UNKNOWN in text
        assert "-0.04R" not in text
        assert "-0.12 USDT" in text              # PnL остаётся доступен

    @pytest.mark.asyncio
    async def test_same_exit_order_id_on_other_symbol_is_not_match(self):
        """Тот же exit orderId на другом инструменте той же сделкой не является."""
        text, _ = await _run_report(
            [_trade(symbol="ETHUSDT", pnl="-3", order_id="CLOSE-TP-1")],
            {},
            exit_evidence={("BTCUSDT", "CLOSE-TP-1"): 3.0},
        )
        assert UNKNOWN in text
        assert "-1R" not in text

    @pytest.mark.asyncio
    async def test_exit_order_id_mismatch_is_unknown(self):
        """Совпал только символ, exit orderId другой → UNKNOWN."""
        text, _ = await _run_report(
            [_trade(symbol="ETHUSDT", pnl="-3", order_id="CLOSE-TP-2")],
            {},
            exit_evidence=_ETH_EXIT_RISK,
        )
        assert UNKNOWN in text
        assert "-1R" not in text

    @pytest.mark.asyncio
    async def test_conflicting_binding_evidence_is_unknown(self):
        """Один (symbol, exit_order_id) с разными рисками → fail-closed UNKNOWN."""
        # Карта строится настоящим journal-хелпером: подставлять сюда готовое
        # число значило бы обойти ровно то правило противоречия, которое
        # проверяется.
        exit_evidence = get_exit_order_risk_evidence([
            _bound(risk=3.0),
            _bound(risk=7.0, entry_order_id="ENTRY-2"),
        ])
        assert exit_evidence == {}
        text, _ = await _run_report(
            [_trade(symbol="ETHUSDT", pnl="-3", order_id="CLOSE-TP-1")],
            {},
            exit_evidence=exit_evidence,
        )
        assert UNKNOWN in text
        assert "-1R" not in text

    @pytest.mark.asyncio
    async def test_legacy_rows_stay_unknown(self):
        """Старые сделки без связи остаются UNKNOWN: backfill запрещён."""
        text, _ = await _run_report(
            [
                _trade(symbol="ETHUSDT", pnl="-3", order_id="OLD-CLOSE-1",
                       ts=1770000000000),
                _trade(symbol="ETHUSDT", pnl="-6", order_id="OLD-CLOSE-2",
                       ts=1770000100000),
            ],
            {},
            exit_evidence=_ETH_EXIT_RISK,
        )
        assert _summary_r(text) == UNKNOWN

    @pytest.mark.asyncio
    async def test_report_never_reads_order_history(self):
        """Ancestry по parentOrderLinkId в runtime отчёта больше не выполняется.

        Production-диагностика доказала, что у SL/TP-детей ``orderLinkId`` и
        ``parentOrderLinkId`` пустые, поэтому месячный scan истории ордеров
        тратил запросы и не давал ни одной связи. Возврат этого запроса —
        регрессия, а не подстраховка.
        """
        calls: list = []
        await _run_report(
            [_trade(symbol="ETHUSDT", pnl="-3", order_id="CLOSE-TP-1")],
            {},
            exit_evidence=_ETH_EXIT_RISK,
            calls=calls,
        )
        assert calls
        assert all(fn is reporting.session.get_closed_pnl for fn, _ in calls)

    def test_ancestry_helpers_are_gone(self):
        """Мёртвый ancestry-код удалён вместе со своей зависимостью в журнале."""
        for name in ("_fetch_close_order_parents", "_validate_history_resp",
                     "_index_history_rows", "get_entry_link_risk_evidence"):
            assert not hasattr(reporting, name)
        import core.journal as journal
        assert not hasattr(journal, "get_entry_link_risk_evidence")
