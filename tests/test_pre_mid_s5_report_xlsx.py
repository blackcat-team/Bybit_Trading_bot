"""
PRE-MID S5 — нативный XLSX месячного отчёта /report MM.YYYY.

Проверяется РЕАЛЬНАЯ книга, которую строит production ``send_report``: документ,
пойманный из ``reply_document``, грузится openpyxl, и ассерты смотрят тип и
числовой формат ячеек, а не только отрендеренный текст. Диспетчер данных
(пагинация, границы месяца, порядок сделок, доказанный исторический R, расчёт
PnL/winrate) не переопределяется — S5 меняет только представление файла.

Доказываемые свойства (контракт задачи A–N):
- A/L: /report MM.YYYY отдаёт валидную .xlsx (не .csv), загружаемую openpyxl,
  с детерминированным именем Report_MM_YYYY.xlsx;
- B: точная первая строка заголовков;
- C: Дата — настоящая datetime-ячейка того же базиса, формат DD.MM.YYYY HH:MM;
- D: Entry/Exit/PnL USDT — числовые ячейки (значения с дробями видят баг
  «строка вместо числа»); PnL несёт знако-цветовой формат, оставаясь числом;
- E: доказанный R — число с сохранённым вычисленным значением;
- F: недоказанный R — ровно тире, не UNKNOWN и не ноль;
- G: источник #prodtest экспортируется как prodtest, встроенные # не трогаются;
- H: внешний текст =1+1 остаётся строкой, а не формулой Excel;
- I: N строк-сделок → N строк данных, без агрегации;
- J: порядок updatedTime-desc сохранён в строках книги;
- K: freeze A2, autofilter A:H по всем строкам, жирный заголовок, вменяемые ширины;
- M: /report без аргумента остаётся текстом и файл не шлёт;
- N: /info говорит про Excel (.xlsx) и не обещает месячный CSV;
- строгая числовая конвертация fail-closed: битое значение рушит файл, а не
  подставляет ноль и не шлёт частичный документ.

Все зависимости Bybit/Telegram замокированы; сетевых вызовов нет.
"""

import io
import os
import sys
from datetime import datetime, timezone
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl

# ── Mock heavy deps before any project import ────────────────────────────────
for _mod in [
    "telegram", "telegram.ext", "telegram.request", "telegram.error",
    "pybit", "pybit.unified_trading",
    "dotenv", "colorama",
]:
    sys.modules.setdefault(_mod, MagicMock())

_UID = "123"

_cfg = MagicMock()
_cfg.ALLOWED_ID = _UID
_cfg.DATA_DIR = Path(__file__).resolve().parent.parent / "data"
sys.modules.setdefault("core.config", _cfg)

for _mod in ["core.trading_core", "core.bybit_call", "core.database", "handlers.orders"]:
    sys.modules.setdefault(_mod, MagicMock())

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

import pytest  # noqa: E402

import handlers.info as info  # noqa: E402
import handlers.reporting as reporting  # noqa: E402
from core.journal import UNKNOWN  # noqa: E402

_HEADERS = [
    "Дата", "Инструмент", "Side", "Entry", "Exit", "PnL USDT", "R", "Источник",
]

# Экспортное имя листа — операторское, единственное.
_SHEET_TITLE = "Сделки"

# Сентябрьские метки времени: отчёт запрашивается за 09.2026, «сейчас» — октябрь,
# поэтому весь месяц попадает в окно выборки без обрезки будущим.
_TS_SEP_02 = 1788370200000   # 2026-09-02 17:30:00 UTC
_TS_SEP_03 = _TS_SEP_02 + 86_400_000
_TS_SEP_04 = _TS_SEP_02 + 2 * 86_400_000

# Preferred ширины колонок из контракта задачи (низ, верх) включительно.
_WIDTH_RANGE = {
    "A": (18, 20), "B": (14, 16), "C": (10, 12), "D": (14, 16),
    "E": (14, 16), "F": (13, 15), "G": (10, 12), "H": (18, 24),
}


class _FixedNow(datetime):
    """datetime с детерминированным now(): «сейчас» — 15 октября 2026."""

    @classmethod
    def now(cls, tz=None):
        return datetime(2026, 10, 15, 12, 0, 0, tzinfo=timezone.utc)


def _trade(symbol="BTCUSDT", *, pnl="7.3", entry="63421.5", exit="63500.25",
           side="Buy", order_id="OID-1", ts=_TS_SEP_02):
    """Строка закрытой сделки в форме ответа get_closed_pnl (с дробями)."""
    return {
        "symbol": symbol, "closedPnl": pnl, "updatedTime": str(ts),
        "side": side, "avgEntryPrice": entry, "avgExitPrice": exit,
        "orderId": order_id,
    }


def _page(trades):
    """Одна терминальная страница ответа Bybit без продолжения."""
    return {"retCode": 0, "retMsg": "OK", "result": {"list": trades}}


async def _run_report(trades, *, evidence=None, exit_evidence=None,
                      sources=None, args=("09.2026",)):
    """Выполняет реальный send_report на замокированных Bybit/Telegram.

    Документ, отданный в reply_document, грузится openpyxl. Возвращает namespace
    с caption, filename, сырыми байтами документа, книгой и текстом последнего
    текстового сообщения — этого хватает всем проверкам S5.
    """
    status_msg = MagicMock()
    status_msg.edit_text = AsyncMock()
    status_msg.delete = AsyncMock()

    update = MagicMock()
    update.effective_user.id = _UID
    update.message.reply_text = AsyncMock(return_value=status_msg)
    update.message.reply_document = AsyncMock()

    context = MagicMock()
    context.args = list(args or [])

    # Сделки отдаются один раз: чанки за месяц не должны их дублировать.
    pages = [_page(list(trades))]

    async def _fake_call(fn, **kw):
        return pages.pop(0) if pages else _page([])

    source_map = dict(sources or {})

    def _fake_source(symbol, ts):
        return source_map.get(symbol, "TG")

    with patch.object(reporting, "ALLOWED_ID", _UID), \
            patch.object(reporting, "datetime", _FixedNow), \
            patch.object(reporting, "bybit_call", new=AsyncMock(side_effect=_fake_call)), \
            patch.object(reporting, "get_source_at_time", new=_fake_source), \
            patch.object(reporting, "get_entry_risk_evidence",
                         return_value=dict(evidence or {})), \
            patch.object(reporting, "get_exit_order_risk_evidence",
                         return_value=dict(exit_evidence or {})), \
            patch("asyncio.sleep", new=AsyncMock()):
        await reporting.send_report(update, context)

    result = SimpleNamespace(
        update=update, status_msg=status_msg,
        reply_text=update.message.reply_text,
        reply_document=update.message.reply_document,
        caption=None, filename=None, document=None, workbook=None, text=None,
    )
    if update.message.reply_document.await_count:
        kwargs = update.message.reply_document.call_args.kwargs
        result.caption = kwargs.get("caption")
        result.filename = kwargs.get("filename")
        document = kwargs.get("document")
        assert isinstance(document, io.BytesIO)
        result.document = document.getvalue()
        document.seek(0)
        result.workbook = openpyxl.load_workbook(document)
    if update.message.reply_text.await_count > 1:
        result.text = update.message.reply_text.call_args_list[-1].args[0]
    return result


def _data_rows(ws):
    """Номера строк с данными (после строки заголовка)."""
    return list(range(2, ws.max_row + 1))


def _rows_by_symbol(ws):
    """Отображение символ → номер строки для сделок."""
    return {ws.cell(row=r, column=2).value: r for r in _data_rows(ws)}


# Отсутствие ключа closedPnl нужно отличать от ключа со значением None.
_MISSING_PNL = object()


def _trade_with_raw_pnl(raw, **kw):
    """Сделка с заданным СЫРЫМ closedPnl; ``_MISSING_PNL`` убирает ключ целиком."""
    trade = _trade(**kw)
    if raw is _MISSING_PNL:
        trade.pop("closedPnl", None)
    else:
        trade["closedPnl"] = raw
    return trade


def _report_error_text(res):
    """Правдивый error UX отчёта из той ветки, что реально сработала.

    Провал до удаления статуса правит статусное сообщение (``edit_text``); провал
    уже после удаления шлёт новое сообщение (``reply_text``). Обе ветки — один и
    тот же truthful error UX, поэтому проверка не должна зависеть от того, на
    каком именно этапе отчёт честно упал.
    """
    if res.reply_text.await_count > 1:
        return res.reply_text.call_args_list[-1].args[0]
    edits = res.status_msg.edit_text.await_args_list
    if edits:
        return edits[-1].args[0]
    return None


# ── A + L + B: реальный загружаемый .xlsx с детерминированным именем ─────────

@pytest.mark.asyncio
async def test_monthly_export_is_real_loadable_xlsx_with_headers():
    """A/L/B: /report 09.2026 → валидная Report_09_2026.xlsx, заголовки точные."""
    res = await _run_report([_trade()], args=["09.2026"])

    assert res.filename == "Report_09_2026.xlsx"
    assert res.filename.endswith(".xlsx")
    assert not res.filename.endswith(".csv")
    # Контейнер OOXML — это ZIP: сигнатура PK, а не CSV/BOM.
    assert res.document[:2] == b"PK"
    assert not res.document.startswith(b"\xef\xbb\xbf")
    # Загружаемость доказана тем, что openpyxl открыл документ в харнессе.
    assert res.workbook is not None
    ws = res.workbook.active
    assert ws.title == _SHEET_TITLE
    assert len(res.workbook.sheetnames) == 1
    assert [c.value for c in ws[1]] == _HEADERS


# ── C: Дата — настоящая datetime-ячейка того же базиса ───────────────────────

@pytest.mark.asyncio
async def test_date_cell_is_true_datetime_with_display_format():
    """C: Дата — datetime (не строка), формат DD.MM.YYYY HH:MM, базис сохранён."""
    res = await _run_report([_trade(ts=_TS_SEP_02)], args=["09.2026"])
    ws = res.workbook.active
    cell = ws.cell(row=2, column=1)

    assert isinstance(cell.value, datetime)
    assert not isinstance(cell.value, str)
    assert cell.number_format == "DD.MM.YYYY HH:MM"
    # Тот же базис интерпретации времени, что в текущем отчёте.
    expected = datetime.fromtimestamp(_TS_SEP_02 / 1000)
    assert abs((cell.value - expected).total_seconds()) < 1


# ── D: Entry/Exit/PnL — настоящие числовые ячейки ────────────────────────────

@pytest.mark.asyncio
async def test_price_and_pnl_cells_are_numeric_not_strings():
    """D: Entry/Exit/PnL USDT — числа с сохранённым значением и своим форматом."""
    res = await _run_report(
        [_trade(entry="63421.5", exit="63500.25", pnl="7.3")], args=["09.2026"],
    )
    ws = res.workbook.active

    entry = ws.cell(row=2, column=4)
    exit_ = ws.cell(row=2, column=5)
    pnl = ws.cell(row=2, column=6)

    for cell, expected in ((entry, 63421.5), (exit_, 63500.25), (pnl, 7.3)):
        assert cell.data_type == "n"
        assert isinstance(cell.value, float)
        assert not isinstance(cell.value, str)
        assert cell.value == expected

    assert entry.number_format == "0.########"
    assert exit_.number_format == "0.########"
    # Знако-цветовой формат PnL: значение остаётся числом.
    assert pnl.number_format == "[Green]+0.00;[Red]-0.00;0.00"


# ── E: доказанный R — число с сохранённым вычисленным значением ──────────────

@pytest.mark.asyncio
async def test_known_r_is_numeric_and_preserves_calculated_value():
    """E: доказанный риск → R числом, ровно pnl/риск, формат 0.00, без литерала R."""
    pnl_raw = "-31.6"
    risk = 5.0
    res = await _run_report(
        [_trade(symbol="GRVTUSDT", pnl=pnl_raw, order_id="OID-G")],
        evidence={("GRVTUSDT", "OID-G"): risk},
        args=["09.2026"],
    )
    ws = res.workbook.active
    r_cell = ws.cell(row=2, column=7)

    assert r_cell.data_type == "n"
    assert isinstance(r_cell.value, float)
    # Идентичное вычисление, что и в production: значение сохранено точно.
    assert r_cell.value == float(pnl_raw) / risk
    assert r_cell.number_format == "0.00"
    # В самой ячейке нет литерала «R».
    assert "R" not in str(r_cell.value)


# ── F: недоказанный R — ровно тире, не UNKNOWN и не ноль ─────────────────────

@pytest.mark.asyncio
async def test_unknown_r_is_em_dash_never_unknown_or_zero():
    """F: риск не доказан → R ровно «—», строкой; не UNKNOWN, не 0, не 0R."""
    res = await _run_report(
        [_trade(symbol="BTCUSDT", pnl="-4.6", order_id="OID-OLD")],
        evidence={},
        args=["09.2026"],
    )
    ws = res.workbook.active
    r_cell = ws.cell(row=2, column=7)

    assert r_cell.value == "—"
    assert r_cell.data_type == "s"
    assert r_cell.value != UNKNOWN
    assert r_cell.value not in (0, 0.0, "0", "0R", "0.00", "0.00R")
    # PnL остаётся доступным числом даже без доказанного R.
    assert ws.cell(row=2, column=6).value == -4.6


# ── G: источник операторский, снимается только ведущий # ────────────────────

@pytest.mark.asyncio
async def test_source_strips_only_leading_hash():
    """G: #prodtest → prodtest; встроенные # сохраняются; данные не мутируются."""
    res = await _run_report(
        [
            _trade(symbol="AAAUSDT", order_id="A", ts=_TS_SEP_02),
            _trade(symbol="BBBUSDT", order_id="B", ts=_TS_SEP_03),
            _trade(symbol="CCCUSDT", order_id="C", ts=_TS_SEP_04),
        ],
        sources={
            "AAAUSDT": "#prodtest",
            "BBBUSDT": "#tg#note",   # ведущий тег снят, встроенный # сохранён
            "CCCUSDT": "telegram",   # без тега — без изменений
        },
        args=["09.2026"],
    )
    ws = res.workbook.active
    by_symbol = _rows_by_symbol(ws)

    def _source(symbol):
        return ws.cell(row=by_symbol[symbol], column=8).value

    assert _source("AAAUSDT") == "prodtest"
    assert _source("BBBUSDT") == "tg#note"
    assert _source("CCCUSDT") == "telegram"
    # Все источники — литеральные строки.
    for symbol in ("AAAUSDT", "BBBUSDT", "CCCUSDT"):
        assert ws.cell(row=by_symbol[symbol], column=8).data_type == "s"


# ── H: внешний текст не становится формулой Excel ────────────────────────────

@pytest.mark.asyncio
async def test_external_text_is_stored_literally_not_as_formula():
    """H: источник =1+1 (и +/-/@) остаётся строкой, а не формулой при открытии."""
    res = await _run_report(
        [
            _trade(symbol="AAAUSDT", order_id="A", ts=_TS_SEP_02),
            _trade(symbol="BBBUSDT", order_id="B", ts=_TS_SEP_03),
        ],
        sources={"AAAUSDT": "=1+1", "BBBUSDT": "@cmd"},
        args=["09.2026"],
    )
    ws = res.workbook.active
    by_symbol = _rows_by_symbol(ws)

    a = ws.cell(row=by_symbol["AAAUSDT"], column=8)
    b = ws.cell(row=by_symbol["BBBUSDT"], column=8)
    assert a.data_type == "s"
    assert a.value == "=1+1"
    assert b.data_type == "s"
    assert b.value == "@cmd"
    # Инструмент и Side тоже литеральные строки.
    assert ws.cell(row=2, column=2).data_type == "s"
    assert ws.cell(row=2, column=3).data_type == "s"


# ── I: одна строка на сделку, без агрегации ──────────────────────────────────

@pytest.mark.asyncio
async def test_one_worksheet_row_per_trade():
    """I: N исходных строк → N строк данных; ничего не схлопывается."""
    trades = [
        _trade(symbol=f"SYM{n}USDT", order_id=f"OID-{n}", pnl=f"{n}.5",
               ts=_TS_SEP_02 + n * 60_000)
        for n in range(1, 5)
    ]
    res = await _run_report(trades, args=["09.2026"])
    ws = res.workbook.active

    assert ws.max_row == len(trades) + 1          # +1 заголовок
    assert len(_data_rows(ws)) == len(trades)
    symbols = {ws.cell(row=r, column=2).value for r in _data_rows(ws)}
    assert symbols == {f"SYM{n}USDT" for n in range(1, 5)}


# ── J: порядок updatedTime-desc сохранён в строках книги ─────────────────────

@pytest.mark.asyncio
async def test_rows_follow_updatedtime_descending_order():
    """J: строки книги идут по убыванию updatedTime, как и текущий отчёт."""
    res = await _run_report(
        [
            _trade(symbol="AAAUSDT", order_id="A", ts=_TS_SEP_02),   # старейшая
            _trade(symbol="BBBUSDT", order_id="B", ts=_TS_SEP_03),
            _trade(symbol="CCCUSDT", order_id="C", ts=_TS_SEP_04),   # новейшая
        ],
        args=["09.2026"],
    )
    ws = res.workbook.active

    order = [ws.cell(row=r, column=2).value for r in _data_rows(ws)]
    assert order == ["CCCUSDT", "BBBUSDT", "AAAUSDT"]
    dates = [ws.cell(row=r, column=1).value for r in _data_rows(ws)]
    assert dates == sorted(dates, reverse=True)


# ── K: нативная Excel-UX ─────────────────────────────────────────────────────

@pytest.mark.asyncio
async def test_workbook_native_excel_ux():
    """K: freeze A2, autofilter A:H по всем строкам, жирный заголовок, ширины."""
    trades = [
        _trade(symbol="AAAUSDT", order_id="A", ts=_TS_SEP_02),
        _trade(symbol="BBBUSDT", order_id="B", ts=_TS_SEP_03),
    ]
    res = await _run_report(trades, args=["09.2026"])
    ws = res.workbook.active

    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == f"A1:H{ws.max_row}"
    # Заголовок жирный по всем восьми колонкам.
    for col in range(1, 9):
        assert ws.cell(row=1, column=col).font.bold is True
    # Вменяемые, не дефолтные ширины в пределах контракта задачи.
    for column, (low, high) in _WIDTH_RANGE.items():
        width = ws.column_dimensions[column].width
        assert width is not None
        assert low <= width <= high


# ── Строгая числовая конвертация fail-closed ────────────────────────────────

@pytest.mark.asyncio
async def test_malformed_numeric_fails_closed_without_document():
    """Битое authoritative-число рушит файл, а не подставляет ноль/частичный XLSX."""
    res = await _run_report(
        [_trade(entry="")],          # пустая цена входа недостоверна
        args=["09.2026"],
    )
    assert res.reply_document.await_count == 0
    assert res.workbook is None
    assert res.text is not None
    assert "Не удалось сформировать отчёт" in res.text


# ── S5-R1: строгая валидация СЫРОГО closedPnl до всех производных ─────────────
# closedPnl — обязательное authoritative-значение. До S5-R1 оно проходило через
# core.utils.safe_float и молча превращалось в 0.0 (True → 1.0) ещё ДО total_pnl,
# wins/losses, R и строки книги: битый источник давал PnL=0 и фальшивый R.
# Теперь сырое значение строго проверяется в единственной точке агрегации, и это
# доказывается через РЕАЛЬНЫЙ send_report, а не прямым вызовом _finite_number.

# Каждое значение здесь должно провалить генерацию отчёта, а не свернуться в ноль.
_INVALID_RAW_PNL = [
    _MISSING_PNL,        # ключ closedPnl отсутствует
    None,                # None
    "",                  # пустая строка
    "   ",               # только пробелы
    "broken",            # мусорная строка
    False,               # bool: раньше → 0.0
    True,                # bool: раньше → 1.0 (и R=0.5 при риске 2)
    [],                  # список
    {},                  # словарь
    float("nan"),        # NaN
    float("inf"),        # +Inf
    float("-inf"),       # -Inf
]


@pytest.mark.asyncio
@pytest.mark.parametrize("raw", _INVALID_RAW_PNL)
async def test_invalid_raw_closed_pnl_fails_report_without_document(raw):
    """Любой недостоверный сырой closedPnl рушит отчёт: ни XLSX, ни PnL=0."""
    res = await _run_report([_trade_with_raw_pnl(raw)], args=["09.2026"])

    # Документ не отправлен и книга не построена: частичного/битого файла нет.
    assert res.reply_document.await_count == 0
    assert res.workbook is None
    # Сработал правдивый error UX отчёта, а не молчаливая подмена нулём.
    err = _report_error_text(res)
    assert err is not None
    assert "Не удалось сформировать отчёт" in err


@pytest.mark.asyncio
@pytest.mark.parametrize("raw", [_MISSING_PNL, None, "", "   ", False, True])
async def test_invalid_raw_closed_pnl_never_fabricates_r_with_proven_risk(raw):
    """Даже при доказанном риске=2 битый closedPnl не даёт документа и R=0/0.5.

    Прямая репродукция QA: False → PnL 0 → R 0; True → PnL 1 → R 0.5;
    missing/None/""/пробелы → PnL 0 → R 0. Проверка сырого значения происходит
    ДО деления на риск, поэтому ни строки книги, ни числового R не возникает.
    """
    res = await _run_report(
        [_trade_with_raw_pnl(raw, symbol="BTCUSDT", order_id="OID-R")],
        evidence={("BTCUSDT", "OID-R"): 2.0},
        args=["09.2026"],
    )

    # Ни одного документа и ни одной строки книги с фабрикованным PnL/R.
    assert res.reply_document.await_count == 0
    assert res.workbook is None
    err = _report_error_text(res)
    assert err is not None
    assert "Не удалось сформировать отчёт" in err


@pytest.mark.asyncio
@pytest.mark.parametrize("zero_raw", [0, 0.0, "0", "0.00"])
async def test_valid_zero_closed_pnl_stays_numeric_with_truthful_zero_r(zero_raw):
    """Легальный ноль — число 0, не ошибка; при доказанном риске R правдиво 0.

    Это отличает валидный ноль от невалидного-в-ноль: книга ОТПРАВЛЯЕТСЯ, PnL —
    числовой 0, ложной ошибки нет.
    """
    res = await _run_report(
        [_trade_with_raw_pnl(zero_raw, symbol="BTCUSDT", order_id="OID-Z")],
        evidence={("BTCUSDT", "OID-Z"): 2.0},
        args=["09.2026"],
    )

    assert res.reply_document.await_count == 1
    ws = res.workbook.active
    pnl_cell = ws.cell(row=2, column=6)
    assert pnl_cell.data_type == "n"
    assert pnl_cell.value == 0
    assert not isinstance(pnl_cell.value, str)
    # Доказанный ненулевой риск + валидный ноль → R числом 0 (0/2), а не ошибка.
    r_cell = ws.cell(row=2, column=7)
    assert r_cell.data_type == "n"
    assert r_cell.value == 0.0


@pytest.mark.asyncio
@pytest.mark.parametrize("raw,expected", [("12.34", 12.34), ("-5.67", -5.67)])
async def test_valid_normal_closed_pnl_stays_numeric_with_proven_r(raw, expected):
    """Обычные значения остаются числовыми ячейками; R по доказанному риску верен."""
    res = await _run_report(
        [_trade_with_raw_pnl(raw, symbol="BTCUSDT", order_id="OID-N")],
        evidence={("BTCUSDT", "OID-N"): 2.0},
        args=["09.2026"],
    )

    assert res.reply_document.await_count == 1
    ws = res.workbook.active
    pnl_cell = ws.cell(row=2, column=6)
    assert pnl_cell.data_type == "n"
    assert pnl_cell.value == expected
    assert not isinstance(pnl_cell.value, str)
    r_cell = ws.cell(row=2, column=7)
    assert r_cell.data_type == "n"
    assert r_cell.value == expected / 2.0


# ── M: /report без аргумента остаётся текстом ────────────────────────────────

@pytest.mark.asyncio
async def test_no_argument_report_stays_text_and_sends_no_file():
    """M: /report без месяца — текстовый отчёт; XLSX/файл не отправляется."""
    res = await _run_report([_trade()], args=[])

    assert res.reply_document.await_count == 0
    assert res.workbook is None
    assert res.text is not None
    assert "Последние 15" in res.text


# ── N: /info описывает Excel/XLSX, а не месячный CSV ─────────────────────────

def test_info_describes_xlsx_not_csv():
    """N: справка /report говорит про Excel (.xlsx) и не обещает CSV за месяц."""
    report_purpose = dict(info.COMMANDS)["/report"]
    assert "xlsx" in report_purpose.lower()
    assert "CSV" not in report_purpose

    message = info.build_info_message(require_market_confirm=True, preview_ttl_sec=300)
    assert "Excel (.xlsx)" in message
    assert "CSV" not in message
